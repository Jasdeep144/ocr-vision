#!/usr/bin/env python3
"""
OCR Web UI — Flask server wrapping ocr_document.py
Run:  python ocr_web.py
Open: http://localhost:5000
"""

import io
import json
import os
import re
import sys
import tempfile
import uuid
from pathlib import Path
from urllib.parse import quote as url_quote

from flask import Flask, jsonify, render_template, request, send_file

sys.path.insert(0, str(Path(__file__).parent))
from ocr_document import make_client, ocr_image_file, ocr_pdf, try_extract_pdf_text, extract_docx_text

app = Flask(__name__, template_folder="html")
app.config["MAX_CONTENT_LENGTH"] = 100 * 1024 * 1024  # 100 MB

# ---------------------------------------------------------------------------
# Document text cache — keyed by session UUID, used for citation look-ups
# ---------------------------------------------------------------------------

DOCUMENT_CACHE: dict[str, dict[str, str]] = {}  # session_id → {filename: full_text}
MAX_CACHE_SESSIONS = 20


def _cache_session(filename_to_text: dict) -> str:
    if len(DOCUMENT_CACHE) >= MAX_CACHE_SESSIONS:
        del DOCUMENT_CACHE[next(iter(DOCUMENT_CACHE))]
    sid = str(uuid.uuid4())
    DOCUMENT_CACHE[sid] = filename_to_text
    return sid


def find_in_text_with_context(text: str, quote: str, context_chars: int = 700) -> dict:
    """Locate a quote in text, tolerating whitespace differences and partial matches.

    Returns {context, highlight_start, highlight_end}.
    highlight_start == -1 means the passage could not be located.
    """
    if not quote or not text:
        return {"context": text[:context_chars * 2], "highlight_start": -1, "highlight_end": -1}

    # Normalize whitespace so OCR newlines inside sentences don't block matching
    norm_text  = re.sub(r'\s+', ' ', text).strip()
    norm_quote = re.sub(r'\s+', ' ', quote).strip()
    lo_norm    = norm_text.lower()

    idx, match_len = -1, len(norm_quote)

    # 1 — exact match on normalized text
    pos = lo_norm.find(norm_quote.lower())
    if pos != -1:
        idx, match_len = pos, len(norm_quote)

    # 2 — sentence-by-sentence on normalized text
    if idx == -1:
        for sent in re.split(r'(?<=[.!?])\s+', norm_quote):
            sent = sent.strip()
            if len(sent) < 20:
                continue
            pos = lo_norm.find(sent.lower())
            if pos != -1:
                idx, match_len = pos, len(sent)
                break

    # 3 — progressive prefix shrinking: 80% → 60% → 40% of normalized quote
    if idx == -1:
        n_words = norm_quote.split()
        for pct in (0.8, 0.6, 0.4):
            prefix = ' '.join(n_words[:max(4, int(len(n_words) * pct))])
            if len(prefix) < 15:
                continue
            pos = lo_norm.find(prefix.lower())
            if pos != -1:
                idx, match_len = pos, len(prefix)
                break

    # 4 — 6-word sliding phrase windows on normalized text
    if idx == -1:
        words = norm_quote.split()
        for start in range(0, max(1, len(words) - 5)):
            phrase = ' '.join(words[start:start + 6])
            if len(phrase) < 15:
                continue
            pos = lo_norm.find(phrase.lower())
            if pos != -1:
                idx, match_len = pos, len(phrase)
                break

    # 5 — token-overlap fallback: find the window with the most shared content words
    if idx == -1:
        q_tokens = {w.lower() for w in norm_quote.split() if len(w) >= 4}
        if q_tokens:
            threshold = max(2, len(q_tokens) // 2)
            t_words   = norm_text.split()
            window    = min(40, max(len(norm_quote.split()) + 10, 20))
            best_score, best_wi = threshold - 1, -1
            for wi in range(max(1, len(t_words) - window + 1)):
                score = len(q_tokens & {w.lower() for w in t_words[wi:wi + window]})
                if score > best_score:
                    best_score, best_wi = score, wi
            if best_wi >= 0:
                char_pos = len(' '.join(t_words[:best_wi])) + (1 if best_wi > 0 else 0)
                idx, match_len = char_pos, len(' '.join(t_words[best_wi:best_wi + window]))

    if idx == -1:
        return {"context": norm_text[:context_chars * 2], "highlight_start": -1, "highlight_end": -1}

    start = max(0, idx - context_chars // 2)
    end   = min(len(norm_text), idx + match_len + context_chars // 2)
    context = norm_text[start:end]
    return {
        "context":         context,
        "highlight_start": idx - start,
        "highlight_end":   idx - start + match_len,
    }


ALLOWED   = {".pdf", ".docx", ".png", ".jpg", ".jpeg", ".tiff", ".tif", ".bmp", ".gif", ".webp"}
IMAGE_EXT = {".png", ".jpg", ".jpeg", ".tiff", ".tif", ".bmp", ".gif", ".webp"}

CHUNK_SIZE           = 60_000   # ~15 k tokens — safe single-call limit for summaries
TRANSLATE_CHUNK_SIZE = 25_000   # ~6 k tokens — tighter limit keeps translation quality high

CHUNK_EXTRACT_PROMPT = (
    "Extract all legally significant content from this document section. Include:\n"
    "• All parties (names, roles, relationships)\n"
    "• All obligations, duties, and commitments\n"
    "• All dates, deadlines, notice periods, and timeframes\n"
    "• All financial terms, payments, fees, and penalties\n"
    "• All rights granted and caps on liability\n"
    "• Termination events and their consequences\n"
    "• Governing law, jurisdiction, and dispute resolution mechanism\n"
    "• Any unusual, onerous, or ambiguous clauses\n\n"
    "Format as structured bullet points. Preserve exact amounts, dates, and names. "
    "Output only the extracted facts — no commentary, no introduction."
)

TRANSLATE_PROMPT = (
    "You are a professional translator. "
    "Translate the following text to English, preserving the original structure, "
    "paragraph breaks, headings, tables, and formatting exactly as they appear. "
    "If the text is already entirely in English, return it completely unchanged. "
    "Output only the translated text — no notes, explanations, or commentary."
)


def chunk_text(text: str, max_chars: int = CHUNK_SIZE) -> list[str]:
    """Split text into chunks at paragraph boundaries, staying under max_chars."""
    if len(text) <= max_chars:
        return [text]
    chunks: list[str] = []
    remaining = text
    while len(remaining) > max_chars:
        pos = remaining.rfind("\n\n", 0, max_chars)
        if pos == -1:
            pos = remaining.rfind("\n", 0, max_chars)
        if pos == -1:
            pos = max_chars
        chunks.append(remaining[:pos].strip())
        remaining = remaining[pos:].strip()
    if remaining:
        chunks.append(remaining)
    return chunks


def prepare_text_for_summary(text: str, client, model: str) -> str:
    """Map-reduce: extract legal facts from each chunk, return condensed extracts.

    If the text fits in a single chunk it is returned unchanged so callers need
    no special-case logic for short documents.
    """
    chunks = chunk_text(text, CHUNK_SIZE)
    if len(chunks) == 1:
        return text
    n = len(chunks)
    extracts: list[str] = []
    for i, chunk in enumerate(chunks, 1):
        prompt = (
            f"This is section {i} of {n} from a legal document.\n\n"
            + CHUNK_EXTRACT_PROMPT
            + "\n\n---\n\n"
            + chunk
        )
        resp = client.messages.create(
            model=model,
            max_tokens=2048,
            messages=[{"role": "user", "content": prompt}],
        )
        extracts.append(f"=== Document Section {i}/{n} ===\n{resp.content[0].text}")
    return "\n\n".join(extracts)

COMBINED_SUMMARY_PROMPT = (
    "You are a senior legal professional reviewing a set of {n} related documents. "
    "Analyse them together as a collection and produce a combined structured legal summary in Markdown.\n\n"
    "Use these exact ## section headers:\n\n"
    "## 1. Document Set Overview\n"
    "Describe what this collection represents as a whole and how the documents relate to each other.\n\n"
    "## 2. Individual Document Summaries\n"
    "Provide a one-paragraph summary for each document.\n\n"
    "## 3. Parties Involved (All Documents)\n"
    "List every party appearing across the set, their roles, and which documents they appear in.\n\n"
    "## 4. Combined Key Terms & Obligations\n"
    "Summarize the core duties and obligations arising from the full document set.\n\n"
    "## 5. Consolidated Dates & Deadlines\n"
    "A unified timeline of all significant dates across all documents.\n\n"
    "## 6. Financial Terms (Consolidated)\n"
    "All payment obligations, amounts, and financial conditions across documents.\n\n"
    "## 7. Cross-Document Issues\n"
    "Identify any inconsistencies, conflicts, gaps, or dependencies between documents. "
    "Flag where documents contradict each other.\n\n"
    "## 8. Governing Law & Jurisdiction\n"
    "Applicable law across all documents; flag any conflicts between jurisdictions.\n\n"
    "## 9. Red Flags & Notable Clauses\n"
    "Critical issues identified across the full document set.\n\n"
    "## 10. Overall Assessment\n"
    "Professional assessment of the document set as a whole — is it consistent, complete, balanced?\n\n"
    "Rules: Use **bold** for key terms. Use bullet points. Write 'Not applicable' where needed. "
    "Be thorough but concise. Output only the formatted summary."
)

LEGAL_SUMMARY_PROMPT = (
    "You are a senior legal professional with expertise in contract review and document analysis. "
    "Analyze the following document and produce a structured legal summary using Markdown formatting.\n\n"
    "Use these exact ## section headers and cover each one:\n\n"
    "## 1. Document Type & Purpose\n"
    "Identify the type of document and its primary legal purpose.\n\n"
    "## 2. Parties Involved\n"
    "List all parties, their roles (e.g. Licensor, Licensee, Employer, Employee), "
    "and any relevant identifying details.\n\n"
    "## 3. Key Terms & Obligations\n"
    "Summarize the core duties, commitments, and obligations of each party.\n\n"
    "## 4. Important Dates & Deadlines\n"
    "List all significant dates: effective date, expiry, notice periods, milestones.\n\n"
    "## 5. Financial Terms\n"
    "Detail payment amounts, schedules, penalties, interest, or any financial conditions.\n\n"
    "## 6. Rights & Liabilities\n"
    "Outline rights granted and any limitations or caps on liability.\n\n"
    "## 7. Termination & Exit Provisions\n"
    "Describe how and when the agreement can be terminated and the consequences.\n\n"
    "## 8. Governing Law & Jurisdiction\n"
    "State the applicable law, jurisdiction, and dispute resolution mechanism.\n\n"
    "## 9. Red Flags & Notable Clauses\n"
    "Highlight any unusual, onerous, ambiguous, or potentially risky provisions "
    "that a lawyer would flag for the client.\n\n"
    "## 10. Overall Assessment\n"
    "Provide a brief professional assessment: is this document balanced, "
    "standard, or does it favour one party? What is the key advice?\n\n"
    "Rules: Use **bold** for key legal terms. Use bullet points for lists. "
    "Write 'Not applicable' for any section that does not apply. "
    "Use precise legal language. Be thorough but concise. "
    "Output only the formatted summary — no preamble or closing remarks."
)


EXCEL_SUMMARY_PROMPT = (
    "You are a senior legal professional. Analyse the document below and return ONLY a valid JSON object. "
    "No markdown, no code fences, no preamble, no commentary — raw JSON only.\n\n"
    "Follow this exact schema. Keep every string value concise (one sentence maximum). "
    'Use "N/A" where information is absent. Arrays must never be null — use [] if nothing applies.\n'
    'For every array item, "source_quote" must be 1-2 sentences copied verbatim from the document '
    'that most directly support that entry. Copy the exact wording — do not paraphrase.\n\n'
    '{\n'
    '  "document_type": "<type and jurisdiction, e.g. Service Agreement — English Law>",\n'
    '  "purpose": "<primary legal purpose in one sentence>",\n'
    '  "effective_date": "<date or Not specified>",\n'
    '  "expiry_date": "<date, Not specified, or Indefinite>",\n'
    '  "parties": [\n'
    '    {"name": "<full legal name>", "role": "<e.g. Licensor, Employer>", "jurisdiction": "<country/state or N/A>", "source_quote": "<verbatim 1-2 sentences>"}\n'
    '  ],\n'
    '  "key_obligations": [\n'
    '    {"party": "<name>", "obligation": "<concise description>", "deadline": "<date, Ongoing, or N/A>", "source_quote": "<verbatim 1-2 sentences>"}\n'
    '  ],\n'
    '  "key_dates": [\n'
    '    {"date": "<date or timeframe>", "event": "<concise event name>", "party": "<responsible party or All Parties>", "source_quote": "<verbatim 1-2 sentences>"}\n'
    '  ],\n'
    '  "financial_terms": [\n'
    '    {"item": "<payment type or fee name>", "amount": "<amount and currency>", "party": "<paying party>", "due_date": "<when due or N/A>", "source_quote": "<verbatim 1-2 sentences>"}\n'
    '  ],\n'
    '  "rights_liabilities": [\n'
    '    {"party": "<name>", "type": "<Right or Liability>", "description": "<concise description>", "cap": "<monetary cap, Uncapped, or N/A>", "source_quote": "<verbatim 1-2 sentences>"}\n'
    '  ],\n'
    '  "termination_provisions": [\n'
    '    {"trigger": "<termination event>", "notice_period": "<e.g. 30 days, Immediate, or N/A>", "consequence": "<key consequence>", "source_quote": "<verbatim 1-2 sentences>"}\n'
    '  ],\n'
    '  "governing_law": {\n'
    '    "jurisdiction": "<country/state>",\n'
    '    "applicable_law": "<e.g. Laws of England and Wales>",\n'
    '    "dispute_resolution": "<e.g. Arbitration — ICC Rules, London>",\n'
    '    "source_quote": "<verbatim 1-2 sentences from the governing law clause>"\n'
    '  },\n'
    '  "red_flags": [\n'
    '    {"severity": "<High or Medium or Low>", "clause": "<clause reference or topic>", "issue": "<concise risk>", "recommendation": "<concise advice>", "source_quote": "<verbatim 1-2 sentences>"}\n'
    '  ],\n'
    '  "overall_assessment": "<2-3 sentence professional assessment of balance, completeness, and key advice>"\n'
    '}\n\n'
    "Return only the JSON object. No markdown. No text before or after."
)


COMBINED_EXCEL_CROSS_DOC_PROMPT = (
    "You are a senior legal professional reviewing {n} documents. "
    "The structured summaries of each document are provided below. "
    "Analyse them together and return ONLY a valid JSON object — no markdown, no code fences, raw JSON only.\n\n"
    "Follow this exact schema. Keep every string concise (one sentence maximum). "
    'Use "N/A" where information is absent. Arrays must never be null — use [] if nothing applies.\n'
    'For each cross-document issue, "citations" must list 1-2 verbatim sentences copied exactly '
    'from the relevant document(s) that evidence the issue. Copy the exact wording — do not paraphrase.\n\n'
    '{\n'
    '  "document_set_overview": "<2-3 sentence description of what this collection represents as a whole>",\n'
    '  "overall_assessment": "<3-4 sentence professional assessment — balance, completeness, consistency, key advice>",\n'
    '  "cross_document_issues": [\n'
    '    {\n'
    '      "severity": "<High or Medium or Low>",\n'
    '      "issue_type": "<e.g. Conflicting Jurisdiction, Overlapping Obligation, Term Mismatch, Missing Counterpart>",\n'
    '      "documents": ["<filename1>", "<filename2>"],\n'
    '      "description": "<concise description of the cross-document issue>",\n'
    '      "recommendation": "<concise advice>",\n'
    '      "citations": [\n'
    '        {"document": "<filename from documents array>", "quote": "<verbatim 1-2 sentences from that document>"}\n'
    '      ]\n'
    '    }\n'
    '  ]\n'
    '}\n\n'
    "Return only the JSON object. No markdown. No text before or after."
)


def _close_truncated_json(raw: str) -> dict:
    """Repair JSON truncated at the token limit by rolling back to the nearest safe cut-point.

    Strategy:
    - Walk the string once, tracking string/escape state and bracket depth.
    - Record every position that is a valid rollback point:
        * after any closing } or ] at any depth
        * after any , at depth >= 1 (between items/properties)
    - Try cut-points from last-to-first (minimal data loss first).
    - For each cut, strip a trailing comma and append the closing characters
      needed to seal all still-open structures, then attempt json.loads.
    - The first attempt that parses cleanly is returned.
    """
    in_str  = False
    esc     = False
    depth   = 0
    stack: list[str]             = []
    safe:  list[tuple[int, list]] = []  # (position, stack snapshot)

    for i, ch in enumerate(raw):
        if esc:
            esc = False
            continue
        if ch == "\\" and in_str:
            esc = True
            continue
        if ch == '"':
            in_str = not in_str
            continue
        if in_str:
            continue

        if ch in ("{", "["):
            depth += 1
            stack.append("}" if ch == "{" else "]")
        elif ch in ("}", "]"):
            depth -= 1
            if stack:
                stack.pop()
            safe.append((i + 1, list(stack)))   # after every closing bracket
        elif ch == "," and depth >= 1:
            safe.append((i, list(stack)))        # after every comma inside structure

    # Try rollback points from last to first (least data loss first)
    for pos, stk in reversed(safe):
        candidate = raw[:pos].rstrip()
        if candidate.endswith(","):
            candidate = candidate[:-1]
        closing = "".join(reversed(stk))
        try:
            return json.loads(candidate + closing)
        except json.JSONDecodeError:
            continue

    raise json.JSONDecodeError("Cannot repair truncated JSON", raw, 0)


def parse_json_response(raw: str) -> dict:
    """Extract and parse a JSON object from a model response.

    Handles four failure modes:
      1. Markdown code fences  — ```json ... ```
      2. Preamble text         — "Here is the analysis:\\n{..."
      3. Postamble text        — "{...}\\nHope this helps!"
      4. Truncated JSON        — response cut at token limit mid-object/array
    """
    raw = raw.strip()

    # Strip markdown code fences
    if raw.startswith("```"):
        raw = raw.split("\n", 1)[1]
        if raw.endswith("```"):
            raw = raw[:-3]
        raw = raw.strip()

    # Find the outermost { ... } block, discarding any surrounding text
    start = raw.find("{")
    end   = raw.rfind("}") + 1
    if start != -1 and end > start:
        raw = raw[start:end]

    # Fast path — well-formed JSON
    try:
        return json.loads(raw)
    except json.JSONDecodeError:
        pass

    # Slow path — attempt to close truncated JSON and retry
    return _close_truncated_json(raw)


def build_combined_excel(docs: list, cross_doc: dict, session_id: str = "") -> bytes:
    """Build a combined multi-document Excel workbook from per-doc JSON + cross-doc analysis."""
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    wb.remove(wb.active)

    # ── Palette ─────────────────────────────────────────────────────────────
    def fill(hex_color):
        return PatternFill("solid", fgColor=hex_color)

    def hdr_font():
        return Font(bold=True, color="FFFFFF", name="Calibri", size=11)

    def bold_font():
        return Font(bold=True, name="Calibri", size=10)

    def body_font():
        return Font(name="Calibri", size=10)

    def wrap():
        return Alignment(wrap_text=True, vertical="top")

    def thin():
        s = Side(style="thin", color="D1D5DB")
        return Border(left=s, right=s, top=s, bottom=s)

    NAVY  = "1E3A5F"
    TEAL  = "0D9488"
    GRAY  = "F1F5F9"
    SEV_FILL = {
        "High":   fill("FEE2E2"),
        "Medium": fill("FEF3C7"),
        "Low":    fill("DCFCE7"),
    }

    def safe(v):
        s = str(v).strip() if v is not None else ""
        return s if s and s.lower() not in ("none", "null") else "N/A"

    def stem(filename):
        return Path(filename).stem

    def auto_widths(ws, max_w=65):
        for col in ws.columns:
            letter = get_column_letter(col[0].column)
            best = max((len(str(c.value or "")) for c in col), default=8)
            ws.column_dimensions[letter].width = min(best + 4, max_w)

    link_font_c = Font(color="0563C1", underline="single", name="Calibri", size=10)

    def cite_url_c(doc_fn, q):
        return (
            f"http://localhost:5000/view-source"
            f"?session={url_quote(session_id)}"
            f"&doc={url_quote(doc_fn)}"
            f"&quote={url_quote(q)}"
        )

    def add_table(title, headers, rows, sev_col=None, quotes=None):
        """quotes: parallel list of (doc_filename, quote_text) | None per row."""
        has_q = quotes and any(q for q in quotes)
        all_headers = list(headers) + (["Source Quote"] if has_q else [])
        ws = wb.create_sheet(title)
        ws.freeze_panes = "A2"
        for ci, h in enumerate(all_headers, 1):
            c = ws.cell(row=1, column=ci, value=h)
            c.fill = fill(NAVY); c.font = hdr_font()
            c.alignment = wrap(); c.border = thin()
        for ri, row in enumerate(rows, 2):
            sev_val  = safe(row[sev_col]) if sev_col is not None and len(row) > sev_col else None
            row_fill = SEV_FILL.get(sev_val) or (fill(GRAY) if ri % 2 == 0 else None)
            for ci, val in enumerate(row, 1):
                c = ws.cell(row=ri, column=ci, value=safe(val))
                if row_fill:
                    c.fill = row_fill
                c.font = body_font(); c.alignment = wrap(); c.border = thin()
            if has_q:
                qt = quotes[ri - 2] if ri - 2 < len(quotes) else None
                qcol = len(all_headers)
                qc = ws.cell(row=ri, column=qcol)
                if qt and qt[1] and session_id:
                    qc.value     = qt[1]
                    qc.hyperlink = cite_url_c(qt[0], qt[1])
                    qc.font      = link_font_c
                else:
                    qc.value = safe(qt[1]) if qt else "N/A"
                    qc.font  = body_font()
                if row_fill:
                    qc.fill = row_fill
                qc.alignment = wrap(); qc.border = thin()
        if ws.max_row > 1:
            ws.auto_filter.ref = ws.dimensions
        auto_widths(ws)
        return ws

    n = len(docs)

    # ── Sheet 1: Document Set Overview ──────────────────────────────────────
    ws_ov = wb.create_sheet("Document Set")
    ws_ov.freeze_panes = "B2"

    tc = ws_ov.cell(row=1, column=1,
                    value=f"Combined Legal Summary  —  {n} Documents")
    ws_ov.merge_cells(f"A1:G1")
    tc.fill = fill(TEAL)
    tc.font = Font(bold=True, color="FFFFFF", name="Calibri", size=13)
    tc.alignment = Alignment(horizontal="center", vertical="center")
    tc.border = thin()
    ws_ov.row_dimensions[1].height = 26

    doc_headers = ["Document", "Type", "Purpose",
                   "Effective Date", "Expiry Date", "Jurisdiction", "Applicable Law"]
    for ci, h in enumerate(doc_headers, 1):
        c = ws_ov.cell(row=2, column=ci, value=h)
        c.fill = fill(NAVY); c.font = hdr_font()
        c.alignment = wrap(); c.border = thin()

    for ri, doc in enumerate(docs, 3):
        d  = doc.get("data", {})
        gl = d.get("governing_law") or {}
        row_fill = fill(GRAY) if ri % 2 == 1 else None
        for ci, val in enumerate([
            stem(doc.get("filename", "document")),
            d.get("document_type"), d.get("purpose"),
            d.get("effective_date"), d.get("expiry_date"),
            gl.get("jurisdiction"), gl.get("applicable_law"),
        ], 1):
            c = ws_ov.cell(row=ri, column=ci, value=safe(val))
            if row_fill:
                c.fill = row_fill
            c.font = body_font(); c.alignment = wrap(); c.border = thin()

    # Overview and Assessment footer block
    next_row = len(docs) + 4
    for label, text_key in [("Document Set Overview", "document_set_overview"),
                             ("Overall Assessment",     "overall_assessment")]:
        lc = ws_ov.cell(row=next_row, column=1, value=label)
        ws_ov.merge_cells(f"A{next_row}:G{next_row}")
        lc.fill = fill(NAVY); lc.font = hdr_font()
        lc.alignment = Alignment(horizontal="left", vertical="center")
        lc.border = thin()
        next_row += 1
        tc2 = ws_ov.cell(row=next_row, column=1,
                          value=safe(cross_doc.get(text_key)))
        ws_ov.merge_cells(f"A{next_row}:G{next_row}")
        tc2.fill = fill("F8FAFC"); tc2.font = body_font()
        tc2.alignment = Alignment(wrap_text=True, vertical="top")
        tc2.border = thin()
        ws_ov.row_dimensions[next_row].height = 60
        next_row += 2

    ws_ov.column_dimensions["A"].width = 22
    ws_ov.column_dimensions["B"].width = 26
    ws_ov.column_dimensions["C"].width = 40
    ws_ov.column_dimensions["D"].width = 14
    ws_ov.column_dimensions["E"].width = 14
    ws_ov.column_dimensions["F"].width = 18
    ws_ov.column_dimensions["G"].width = 28

    # ── Sheets 2-7: Aggregated per-document data ─────────────────────────────
    def agg(field, mapper):
        rows, quotes = [], []
        for doc in docs:
            d   = doc.get("data", {})
            fn  = doc.get("filename", "document")
            src = stem(fn)
            for item in (d.get(field) or []):
                rows.append([src] + mapper(item))
                quotes.append((fn, item.get("source_quote") or ""))
        return rows, quotes

    _p_rows, _p_q = agg("parties", lambda p: [p.get("name"), p.get("role"), p.get("jurisdiction")])
    add_table("Parties",
        ["Source Document", "Name", "Role", "Jurisdiction"],
        _p_rows, quotes=_p_q)

    _o_rows, _o_q = agg("key_obligations", lambda o: [o.get("party"), o.get("obligation"), o.get("deadline")])
    add_table("Obligations",
        ["Source Document", "Party", "Obligation", "Deadline"],
        _o_rows, quotes=_o_q)

    _d_rows, _d_q = agg("key_dates", lambda x: [x.get("date"), x.get("event"), x.get("party")])
    add_table("Key Dates",
        ["Source Document", "Date", "Event", "Party"],
        _d_rows, quotes=_d_q)

    _f_rows, _f_q = agg("financial_terms",
        lambda f: [f.get("item"), f.get("amount"), f.get("party"), f.get("due_date")])
    add_table("Financial Terms",
        ["Source Document", "Item", "Amount", "Party", "Due Date"],
        _f_rows, quotes=_f_q)

    _rl_rows, _rl_q = agg("rights_liabilities",
        lambda r: [r.get("party"), r.get("type"), r.get("description"), r.get("cap")])
    add_table("Rights & Liabilities",
        ["Source Document", "Party", "Type", "Description", "Cap / Limit"],
        _rl_rows, quotes=_rl_q)

    _t_rows, _t_q = agg("termination_provisions",
        lambda t: [t.get("trigger"), t.get("notice_period"), t.get("consequence")])
    add_table("Termination",
        ["Source Document", "Trigger", "Notice Period", "Consequence"],
        _t_rows, quotes=_t_q)

    # ── Sheet 8: Cross-Document Issues ───────────────────────────────────────
    cross_issues = cross_doc.get("cross_document_issues") or []
    cross_rows = [
        [i.get("severity"), i.get("issue_type"),
         ", ".join(i.get("documents") or []),
         i.get("description"), i.get("recommendation")]
        for i in cross_issues
    ]
    cross_quotes = []
    for i in cross_issues:
        cits = i.get("citations") or []
        if cits:
            cross_quotes.append((cits[0].get("document", ""), cits[0].get("quote", "")))
        else:
            cross_quotes.append(None)
    add_table("Cross-Doc Issues",
        ["Severity", "Issue Type", "Documents", "Description", "Recommendation"],
        cross_rows, sev_col=0, quotes=cross_quotes)

    # ── Sheet 9: Red Flags (all docs) ────────────────────────────────────────
    rf_rows, rf_quotes = [], []
    for doc in docs:
        fn  = doc.get("filename", "document")
        src = stem(fn)
        for r in (doc.get("data", {}).get("red_flags") or []):
            rf_rows.append([src, r.get("severity"), r.get("clause"),
                            r.get("issue"), r.get("recommendation")])
            rf_quotes.append((fn, r.get("source_quote") or ""))
    add_table("Red Flags",
        ["Source Document", "Severity", "Clause / Topic", "Issue", "Recommendation"],
        rf_rows, sev_col=1, quotes=rf_quotes)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_excel_summary(data: dict, doc_filename: str, session_id: str = "") -> bytes:
    """Build a styled multi-sheet Excel workbook from the structured legal summary JSON."""
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    wb.remove(wb.active)  # remove default blank sheet

    # ── Palette ──────────────────────────────────────────────────────────────
    def fill(hex_color):
        return PatternFill("solid", fgColor=hex_color)

    def hdr_font():
        return Font(bold=True, color="FFFFFF", name="Calibri", size=11)

    def bold_font():
        return Font(bold=True, name="Calibri", size=10)

    def body_font():
        return Font(name="Calibri", size=10)

    def wrap():
        return Alignment(wrap_text=True, vertical="top")

    def thin():
        s = Side(style="thin", color="D1D5DB")
        return Border(left=s, right=s, top=s, bottom=s)

    NAVY    = "1E3A5F"
    TEAL    = "0D9488"
    GRAY    = "F1F5F9"
    SEV_FILL = {
        "High":   fill("FEE2E2"),
        "Medium": fill("FEF3C7"),
        "Low":    fill("DCFCE7"),
    }

    def safe(v):
        s = str(v).strip() if v is not None else ""
        return s if s and s.lower() not in ("none", "null") else "N/A"

    def auto_widths(ws, max_w=65):
        for col in ws.columns:
            letter = get_column_letter(col[0].column)
            best = max((len(str(c.value or "")) for c in col), default=8)
            ws.column_dimensions[letter].width = min(best + 4, max_w)

    link_font = Font(color="0563C1", underline="single", name="Calibri", size=10)

    def cite_url(doc_fn, q):
        return (
            f"http://localhost:5000/view-source"
            f"?session={url_quote(session_id)}"
            f"&doc={url_quote(doc_fn)}"
            f"&quote={url_quote(q)}"
        )

    # ── Reusable table-sheet builder ─────────────────────────────────────────
    def add_table(title, headers, rows, sev_col=None, quotes=None):
        """quotes: parallel list of (doc_filename, quote_text) | None per row."""
        has_q = quotes and any(q for q in quotes)
        all_headers = list(headers) + (["Source Quote"] if has_q else [])
        ws = wb.create_sheet(title)
        ws.freeze_panes = "A2"
        # Header row
        for ci, h in enumerate(all_headers, 1):
            c = ws.cell(row=1, column=ci, value=h)
            c.fill = fill(NAVY)
            c.font = hdr_font()
            c.alignment = wrap()
            c.border = thin()
        # Data rows
        for ri, row in enumerate(rows, 2):
            sev_val  = safe(row[sev_col]) if sev_col is not None and len(row) > sev_col else None
            row_fill = SEV_FILL.get(sev_val) or (fill(GRAY) if ri % 2 == 0 else None)
            for ci, val in enumerate(row, 1):
                c = ws.cell(row=ri, column=ci, value=safe(val))
                if row_fill:
                    c.fill = row_fill
                c.font      = body_font()
                c.alignment = wrap()
                c.border    = thin()
            if has_q:
                qt = quotes[ri - 2] if ri - 2 < len(quotes) else None
                qcol = len(all_headers)
                qc = ws.cell(row=ri, column=qcol)
                if qt and qt[1] and session_id:
                    qc.value     = qt[1]
                    qc.hyperlink = cite_url(qt[0], qt[1])
                    qc.font      = link_font
                else:
                    qc.value = safe(qt[1]) if qt else "N/A"
                    qc.font  = body_font()
                if row_fill:
                    qc.fill = row_fill
                qc.alignment = wrap()
                qc.border    = thin()
        if ws.max_row > 1:
            ws.auto_filter.ref = ws.dimensions
        auto_widths(ws)
        return ws

    # ── Sheet 1: Overview (key-value) ────────────────────────────────────────
    ws_ov = wb.create_sheet("Overview")
    ws_ov.freeze_panes = "B2"

    # Title banner
    title_val = f"Legal Summary  —  {Path(doc_filename).stem}"
    tc = ws_ov.cell(row=1, column=1, value=title_val)
    ws_ov.merge_cells("A1:B1")
    tc.fill      = fill(TEAL)
    tc.font      = Font(bold=True, color="FFFFFF", name="Calibri", size=13)
    tc.alignment = Alignment(horizontal="center", vertical="center")
    tc.border    = thin()
    ws_ov.row_dimensions[1].height = 26

    gl = data.get("governing_law") or {}
    overview_rows = [
        ("Document Type",      data.get("document_type")),
        ("Purpose",            data.get("purpose")),
        ("Effective Date",     data.get("effective_date")),
        ("Expiry Date",        data.get("expiry_date")),
        ("Jurisdiction",       gl.get("jurisdiction")),
        ("Applicable Law",     gl.get("applicable_law")),
        ("Dispute Resolution", gl.get("dispute_resolution")),
        ("Overall Assessment", data.get("overall_assessment")),
    ]
    for ri, (field, value) in enumerate(overview_rows, 2):
        alt = fill(GRAY) if ri % 2 == 0 else fill("FFFFFF")
        fc = ws_ov.cell(row=ri, column=1, value=field)
        vc = ws_ov.cell(row=ri, column=2, value=safe(value))
        for c in (fc, vc):
            c.fill = alt; c.alignment = wrap(); c.border = thin()
        fc.font = bold_font()
        vc.font = body_font()
    ws_ov.column_dimensions["A"].width = 22
    ws_ov.column_dimensions["B"].width = 82

    fn = doc_filename

    def sq(items, key="source_quote"):
        return [(fn, i.get(key) or "") for i in (items or [])]

    # ── Sheet 2: Parties ─────────────────────────────────────────────────────
    _parties = data.get("parties") or []
    add_table("Parties",
        ["Name", "Role", "Jurisdiction"],
        [[p.get("name"), p.get("role"), p.get("jurisdiction")] for p in _parties],
        quotes=sq(_parties))

    # ── Sheet 3: Key Obligations ─────────────────────────────────────────────
    _obls = data.get("key_obligations") or []
    add_table("Obligations",
        ["Party", "Obligation", "Deadline"],
        [[o.get("party"), o.get("obligation"), o.get("deadline")] for o in _obls],
        quotes=sq(_obls))

    # ── Sheet 4: Key Dates ───────────────────────────────────────────────────
    _dates = data.get("key_dates") or []
    add_table("Key Dates",
        ["Date", "Event", "Party"],
        [[d.get("date"), d.get("event"), d.get("party")] for d in _dates],
        quotes=sq(_dates))

    # ── Sheet 5: Financial Terms ─────────────────────────────────────────────
    _fin = data.get("financial_terms") or []
    add_table("Financial Terms",
        ["Item", "Amount", "Party", "Due Date"],
        [[f.get("item"), f.get("amount"), f.get("party"), f.get("due_date")] for f in _fin],
        quotes=sq(_fin))

    # ── Sheet 6: Rights & Liabilities ───────────────────────────────────────
    _rl = data.get("rights_liabilities") or []
    add_table("Rights & Liabilities",
        ["Party", "Type", "Description", "Cap / Limit"],
        [[r.get("party"), r.get("type"), r.get("description"), r.get("cap")] for r in _rl],
        quotes=sq(_rl))

    # ── Sheet 7: Termination ─────────────────────────────────────────────────
    _term = data.get("termination_provisions") or []
    add_table("Termination",
        ["Trigger", "Notice Period", "Consequence"],
        [[t.get("trigger"), t.get("notice_period"), t.get("consequence")] for t in _term],
        quotes=sq(_term))

    # ── Sheet 8: Red Flags (severity-coloured) ───────────────────────────────
    _rf = data.get("red_flags") or []
    add_table("Red Flags",
        ["Severity", "Clause / Topic", "Issue", "Recommendation"],
        [[r.get("severity"), r.get("clause"), r.get("issue"), r.get("recommendation")] for r in _rf],
        sev_col=0, quotes=sq(_rf))

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


@app.route("/")
def index():
    return render_template("index.html")


@app.route("/api/ocr", methods=["POST"])
def ocr_endpoint():
    if "file" not in request.files:
        return jsonify({"error": "No file provided"}), 400

    file = request.files["file"]
    if not file.filename:
        return jsonify({"error": "No file selected"}), 400

    suffix = Path(file.filename).suffix.lower()
    if suffix not in ALLOWED:
        return jsonify({"error": f'Unsupported file type "{suffix}"'}), 400

    api_key = os.environ.get("ANTHROPIC_API_KEY", "").strip()
    if not api_key:
        return jsonify({
            "error": "ANTHROPIC_API_KEY is not set. "
                     "Add it to your environment and restart the server."
        }), 400

    model = request.form.get("model", "claude-sonnet-4-6")
    force_images = request.form.get("force_images", "false").lower() == "true"
    skip_direct  = request.form.get("skip_direct",  "false").lower() == "true"
    dpi = int(request.form.get("dpi", 200))

    client = make_client(api_key)

    with tempfile.NamedTemporaryFile(suffix=suffix, delete=False) as tmp:
        file.save(tmp.name)
        tmp_path = Path(tmp.name)

    try:
        if suffix == ".docx":
            text = extract_docx_text(tmp_path)
            if not text:
                return jsonify({
                    "error": "Could not extract text from this Word document. "
                             "It may contain only embedded images — save as PDF and retry."
                }), 400
            return jsonify({"text": text, "filename": file.filename, "method": "direct"})

        elif suffix == ".pdf":
            if not force_images and not skip_direct:
                direct_text = try_extract_pdf_text(tmp_path)
                if direct_text:
                    return jsonify({"text": direct_text, "filename": file.filename, "method": "direct"})
            text = ocr_pdf(tmp_path, client, model, dpi, force_images)
            return jsonify({"text": text, "filename": file.filename, "method": "ocr"})

        else:
            text = ocr_image_file(tmp_path, client, model)
            return jsonify({"text": text, "filename": file.filename, "method": "ocr"})

    except Exception as exc:
        return jsonify({"error": str(exc)}), 500
    finally:
        tmp_path.unlink(missing_ok=True)


@app.route("/api/translate", methods=["POST"])
def translate_endpoint():
    data = request.get_json()
    text = (data or {}).get("text", "").strip()
    if not text:
        return jsonify({"error": "No text provided"}), 400

    api_key = os.environ.get("ANTHROPIC_API_KEY", "").strip()
    if not api_key:
        return jsonify({"error": "ANTHROPIC_API_KEY is not set"}), 400

    model = (data or {}).get("model", "claude-sonnet-4-6")
    client = make_client(api_key)

    try:
        chunks = chunk_text(text, TRANSLATE_CHUNK_SIZE)
        parts: list[str] = []
        for chunk in chunks:
            resp = client.messages.create(
                model=model,
                max_tokens=8096,
                messages=[{"role": "user", "content": TRANSLATE_PROMPT + "\n\n---\n\n" + chunk}],
            )
            parts.append(resp.content[0].text)
        return jsonify({"text": "\n\n".join(parts)})
    except Exception as exc:
        return jsonify({"error": str(exc)}), 500


@app.route("/api/legal-summary", methods=["POST"])
def legal_summary_endpoint():
    data = request.get_json()
    text = (data or {}).get("text", "").strip()
    if not text:
        return jsonify({"error": "No text provided"}), 400

    api_key = os.environ.get("ANTHROPIC_API_KEY", "").strip()
    if not api_key:
        return jsonify({"error": "ANTHROPIC_API_KEY is not set"}), 400

    model = (data or {}).get("model", "claude-sonnet-4-6")
    client = make_client(api_key)

    try:
        condensed = prepare_text_for_summary(text, client, model)
        response = client.messages.create(
            model=model,
            max_tokens=8192,
            messages=[{"role": "user", "content": LEGAL_SUMMARY_PROMPT + "\n\n---\n\n" + condensed}],
        )
        return jsonify({"summary": response.content[0].text})
    except Exception as exc:
        return jsonify({"error": str(exc)}), 500


@app.route("/api/combined-summary", methods=["POST"])
def combined_summary_endpoint():
    data = request.get_json()
    docs = (data or {}).get("documents", [])  # [{filename, text}, ...]
    if not docs or len(docs) < 2:
        return jsonify({"error": "At least 2 documents are required"}), 400

    api_key = os.environ.get("ANTHROPIC_API_KEY", "").strip()
    if not api_key:
        return jsonify({"error": "ANTHROPIC_API_KEY is not set"}), 400

    model = (data or {}).get("model", "claude-sonnet-4-6")
    client = make_client(api_key)

    try:
        condensed_parts: list[str] = []
        for i, doc in enumerate(docs, 1):
            doc_text = doc.get("text", "").strip()
            condensed = prepare_text_for_summary(doc_text, client, model)
            condensed_parts.append(f"--- Document {i}: {doc['filename']} ---\n\n{condensed}")

        combined = "\n\n".join(condensed_parts)
        prompt = COMBINED_SUMMARY_PROMPT.format(n=len(docs)) + "\n\n" + combined

        response = client.messages.create(
            model=model,
            max_tokens=8096,
            messages=[{"role": "user", "content": prompt}],
        )
        return jsonify({"summary": response.content[0].text})
    except Exception as exc:
        return jsonify({"error": str(exc)}), 500


@app.route("/api/excel-summary", methods=["POST"])
def excel_summary_endpoint():
    data     = request.get_json()
    text     = (data or {}).get("text", "").strip()
    filename = (data or {}).get("filename", "document")
    if not text:
        return jsonify({"error": "No text provided"}), 400

    api_key = os.environ.get("ANTHROPIC_API_KEY", "").strip()
    if not api_key:
        return jsonify({"error": "ANTHROPIC_API_KEY is not set"}), 400

    model  = (data or {}).get("model", "claude-sonnet-4-6")
    client = make_client(api_key)

    try:
        condensed = prepare_text_for_summary(text, client, model)
        response = client.messages.create(
            model=model,
            max_tokens=8192,
            messages=[{"role": "user", "content": EXCEL_SUMMARY_PROMPT + "\n\n---\n\n" + condensed}],
        )
        summary_data = parse_json_response(response.content[0].text)
        excel_bytes  = build_excel_summary(summary_data, filename)

        stem    = Path(filename).stem
        dl_name = f"{stem}_legal_summary.xlsx"
        return send_file(
            io.BytesIO(excel_bytes),
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=dl_name,
        )
    except json.JSONDecodeError as e:
        return jsonify({"error": f"Failed to parse structured response: {e}"}), 500
    except Exception as exc:
        return jsonify({"error": str(exc)}), 500


@app.route("/api/combined-legal-data", methods=["POST"])
def combined_legal_data_endpoint():
    """Two-pass extraction: per-doc structured JSON + cross-doc synthesis."""
    data = request.get_json()
    docs = (data or {}).get("documents", [])
    if not docs or len(docs) < 2:
        return jsonify({"error": "At least 2 documents are required"}), 400

    api_key = os.environ.get("ANTHROPIC_API_KEY", "").strip()
    if not api_key:
        return jsonify({"error": "ANTHROPIC_API_KEY is not set"}), 400

    model  = (data or {}).get("model", "claude-sonnet-4-6")
    client = make_client(api_key)

    try:
        text_cache = {doc["filename"]: doc.get("text", "") for doc in docs}
        session_id = _cache_session(text_cache)

        # Pass 1 — structured JSON for each document
        per_doc: list[dict] = []
        for doc in docs:
            text     = doc.get("text", "").strip()
            filename = doc.get("filename", "document")
            condensed = prepare_text_for_summary(text, client, model)
            text_cache[f"{filename}::condensed"] = condensed
            resp = client.messages.create(
                model=model, max_tokens=8192,
                messages=[{"role": "user",
                           "content": EXCEL_SUMMARY_PROMPT + "\n\n---\n\n" + condensed}],
            )
            per_doc.append({"filename": filename,
                            "data": parse_json_response(resp.content[0].text)})

        # Pass 2 — cross-document synthesis
        summaries = "\n\n".join(
            f"--- Document {i+1}: {r['filename']} ---\n{json.dumps(r['data'], indent=2)}"
            for i, r in enumerate(per_doc)
        )
        cross_resp = client.messages.create(
            model=model, max_tokens=8192,
            messages=[{"role": "user",
                       "content": COMBINED_EXCEL_CROSS_DOC_PROMPT.replace("{n}", str(len(docs)))
                                  + "\n\n" + summaries}],
        )
        cross_doc = parse_json_response(cross_resp.content[0].text)
        return jsonify({"docs": per_doc, "cross_doc": cross_doc, "session_id": session_id})

    except json.JSONDecodeError as e:
        return jsonify({"error": f"Failed to parse structured response: {e}"}), 500
    except Exception as exc:
        return jsonify({"error": str(exc)}), 500


@app.route("/api/combined-excel-from-json", methods=["POST"])
def combined_excel_from_json_endpoint():
    """Build combined xlsx from cached JSON — no Claude call."""
    body       = request.get_json()
    docs       = (body or {}).get("docs")
    cross_doc  = (body or {}).get("cross_doc")
    session_id = (body or {}).get("session_id", "")
    if not docs or not cross_doc:
        return jsonify({"error": "Missing docs or cross_doc"}), 400
    try:
        excel_bytes = build_combined_excel(docs, cross_doc, session_id)
        return send_file(
            io.BytesIO(excel_bytes),
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name="combined_legal_summary.xlsx",
        )
    except Exception as exc:
        return jsonify({"error": str(exc)}), 500


@app.route("/api/legal-data", methods=["POST"])
def legal_data_endpoint():
    """Return the structured legal JSON (same schema as excel-summary) without building the xlsx."""
    data     = request.get_json()
    text     = (data or {}).get("text", "").strip()
    filename = (data or {}).get("filename", "document")
    if not text:
        return jsonify({"error": "No text provided"}), 400

    api_key = os.environ.get("ANTHROPIC_API_KEY", "").strip()
    if not api_key:
        return jsonify({"error": "ANTHROPIC_API_KEY is not set"}), 400

    model  = (data or {}).get("model", "claude-sonnet-4-6")
    client = make_client(api_key)

    try:
        text_cache = {filename: text}
        session_id = _cache_session(text_cache)
        condensed  = prepare_text_for_summary(text, client, model)
        text_cache[f"{filename}::condensed"] = condensed
        response = client.messages.create(
            model=model,
            max_tokens=8192,
            messages=[{"role": "user", "content": EXCEL_SUMMARY_PROMPT + "\n\n---\n\n" + condensed}],
        )
        summary_data = parse_json_response(response.content[0].text)
        return jsonify({"data": summary_data, "filename": filename, "session_id": session_id})
    except json.JSONDecodeError as e:
        return jsonify({"error": f"Failed to parse structured response: {e}"}), 500
    except Exception as exc:
        return jsonify({"error": str(exc)}), 500


@app.route("/api/excel-from-json", methods=["POST"])
def excel_from_json_endpoint():
    """Build and return an xlsx from already-structured JSON — no Claude call needed."""
    body         = request.get_json()
    summary_data = (body or {}).get("data")
    filename     = (body or {}).get("filename", "document")
    session_id   = (body or {}).get("session_id", "")
    if not summary_data:
        return jsonify({"error": "No data provided"}), 400

    try:
        excel_bytes = build_excel_summary(summary_data, filename, session_id)
        stem    = Path(filename).stem
        dl_name = f"{stem}_legal_summary.xlsx"
        return send_file(
            io.BytesIO(excel_bytes),
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=dl_name,
        )
    except Exception as exc:
        return jsonify({"error": str(exc)}), 500


@app.route("/api/source-text")
def source_text_endpoint():
    """Return context around a cited passage for the web-viewer citation panel."""
    session_id = request.args.get("session", "")
    filename   = request.args.get("doc", "")
    quote      = request.args.get("quote", "")

    session_docs = DOCUMENT_CACHE.get(session_id)
    if not session_docs:
        return jsonify({"error": "Session not found or expired — re-run the analysis"}), 404

    text = session_docs.get(filename, "")
    if not text:
        for fn, t in session_docs.items():
            if filename in fn or fn in filename:
                text = t
                filename = fn
                break
    if not text:
        return jsonify({"error": f"Document '{filename}' not found in session"}), 404

    result = find_in_text_with_context(text, quote)
    if result["highlight_start"] == -1:
        condensed_text = session_docs.get(f"{filename}::condensed", "")
        if condensed_text:
            r2 = find_in_text_with_context(condensed_text, quote)
            if r2["highlight_start"] != -1:
                r2["filename"] = filename
                return jsonify(r2)
    result["filename"] = filename
    return jsonify(result)


@app.route("/view-source")
def view_source_page():
    """Simple HTML page showing cited passage with highlighting — used by .xlsx hyperlinks."""
    session_id = request.args.get("session", "")
    filename   = request.args.get("doc", "")
    quote      = request.args.get("quote", "")

    session_docs = DOCUMENT_CACHE.get(session_id)
    if not session_docs:
        return (
            "<!doctype html><html><body style='font-family:sans-serif;padding:40px;color:#374151'>"
            "<h2 style='color:#EF4444'>Session Expired</h2>"
            "<p>The citation session has expired (server was restarted).</p>"
            "<p>Re-run the Excel analysis in the web app to generate fresh citation links.</p>"
            "<p style='margin-top:20px;color:#6B7280;font-size:13px'>"
            f"<strong>Document:</strong> {filename}<br>"
            f"<strong>Quote:</strong> {quote}"
            "</p></body></html>",
            404,
        )

    text = session_docs.get(filename, "")
    if not text:
        for fn, t in session_docs.items():
            if filename in fn or fn in filename:
                text, filename = t, fn
                break
    if not text:
        return (
            "<!doctype html><html><body style='font-family:sans-serif;padding:40px;color:#374151'>"
            f"<h2 style='color:#EF4444'>Document Not Found</h2>"
            f"<p>'{filename}' is not in this session.</p></body></html>",
            404,
        )

    r = find_in_text_with_context(text, quote)
    if r["highlight_start"] == -1:
        condensed_text = session_docs.get(f"{filename}::condensed", "")
        if condensed_text:
            r2 = find_in_text_with_context(condensed_text, quote)
            if r2["highlight_start"] != -1:
                r = r2
    ctx = r["context"]
    hs, he = r["highlight_start"], r["highlight_end"]

    import html as html_mod
    if hs >= 0:
        body_html = (
            html_mod.escape(ctx[:hs])
            + f'<mark style="background:#FEF08A;border-radius:2px;padding:0 2px">{html_mod.escape(ctx[hs:he])}</mark>'
            + html_mod.escape(ctx[he:])
        )
        note = ""
    else:
        body_html = html_mod.escape(ctx)
        note = "<p style='color:#94A3B8;font-size:12px;margin-bottom:12px'>[Exact location not found — showing document excerpt]</p>"

    page = f"""<!doctype html>
<html>
<head>
  <meta charset="utf-8">
  <title>Source: {html_mod.escape(filename)}</title>
  <style>
    body {{font-family: Calibri, sans-serif; padding: 32px 48px; max-width: 860px;
           margin: 0 auto; color: #374151; background: #F9FAFB}}
    h2   {{font-size: 15px; color: #1E3A5F; margin-bottom: 4px}}
    .doc {{font-size: 12px; color: #6B7280; margin-bottom: 20px}}
    pre  {{white-space: pre-wrap; font-family: inherit; font-size: 13px;
           line-height: 1.75; background: #fff; border: 1px solid #E2E8F0;
           border-radius: 6px; padding: 20px 24px}}
  </style>
</head>
<body>
  <h2>Source Reference</h2>
  <p class="doc">{html_mod.escape(filename)}</p>
  {note}
  <pre>{body_html}</pre>
</body>
</html>"""
    return page


if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    print("━" * 48)
    print("  OCR Vision UI")
    print(f"  ➜  http://localhost:{port}")
    print("━" * 48)
    app.run(debug=False, host="0.0.0.0", port=port)
